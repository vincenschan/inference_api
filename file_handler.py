# -*- coding: utf-8 -*-
# __author: vincens
# @file: handle_base.py
# @time: 2022 07 30
# @email: vincenschan@163.com
# Ref: https://www.jianshu.com/p/f0256daaf9ab

import csv
import json
import logging
import os
import re

# from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE
# import xlsxwriter as xw
import openpyxl

ILLEGAL_CHARACTERS_RE = re.compile(r'[\000-\010]|[\013-\014]|[\016-\037]')


class BaseHandler(object):
    def __init__(self, logger=logging):
        self.logger = logger

    def read(self, *args, **kwargs):
        pass

    def write(self, *args, **kwargs):
        pass

    def append(self, *args, **kwargs):
        pass

    @classmethod
    def remove_path(cls, path):
        """
        删除文件 或者 文件夹
        :param path:
        :return:
        """
        os.remove(path)

    @classmethod
    def make_path(cls, path):
        """
        创建新的文件夹路径
        :param path:
        :return:
        """

        if not os.path.exists(path):
            os.makedirs(path)
        pass

    @classmethod
    def check_suffix(cls, path, check):
        """

        :param path:
        :param check:
        :return:
        """

        if os.path.splitext(path)[-1][1:].upper() != check.upper():
            return False
        return True

    @classmethod
    def get_file_name(cls, path):
        """

        :param path:
        :return:
        """
        return os.path.splitext(path)[-1]

    @classmethod
    def get_file_dir(cls, path):
        """

        :param path:
        :return:
        """
        return os.path.splitext(path)[0]


class ConvertHandler(object):
    def __init__(self):
        pass

    def to_txt(self, *args, **kwargs):
        pass

    def to_excel(self, *args, **kwargs):
        pass

    def to_json(self, *args, **kwargs):
        pass

    def list_converted_dictionary(self, data, key_index, value_indexs=None):
        """

        :param data:
        :param key_index:
        :param value_indexs:
        :return:
        """

        conv_data = {}
        for i, contents in enumerate(data):
            _key = data[key_index]
            conv_contents = []
            if value_indexs is not None:
                for _index in value_indexs:
                    conv_contents.append(contents[_index])
            if _key and conv_contents:
                if _key not in conv_data:
                    conv_data.update({_key: [conv_contents]})
                else:
                    conv_data[_key].append(conv_contents)
        return conv_data

    def dictionary_converted_list(self, data, key_indexs):
        """

        :param data:
        :param key_indexs:
        :return:
        """
        res = []
        for i, _key in enumerate(data.keys()):
            contents = data[_key]
            tmp_cont = []
            for j, cont in enumerate(contents):
                if j in key_indexs:
                    tmp_cont.append(cont)
            res.append(tmp_cont[:])
        return res


class CSVHandler(BaseHandler, ConvertHandler):
    """
        r	以只读方式打开文件。文件的指针将会放在文件的开头。这是默认模式。
        r+	打开一个文件用于读写。文件指针将会放在文件的开头。
        w+	打开一个文件用于读写。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。
        a	打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。也就是说，新的内容将会被写入到已有内容之后。如果该文件不存在，创建新文件进行写入。
        a+	打开一个文件用于读写。如果该文件已存在，文件指针将会放在文件的结尾。文件打开时会是追加模式。如果该文件不存在，创建新文件用于读写。
    """

    def __init__(self):
        super(CSVHandler, self).__init__()

    def read(self, csv_path, mode="r", encoding="utf-8-sig"):
        """

        :param csv_path:
        :param mode:
        :param encoding:
        :return:
        """
        csv_datas = []
        try:
            with open(csv_path, mode=mode, encoding=encoding) as f:
                reader = csv.reader((line.replace('\0', '') for line in f), delimiter=",")
                header = next(reader)
                # 逐行获取数据，并输出
                for row in reader:
                    csv_datas.append(row)
            return header, csv_datas
        except Exception as e:
            self.logger.error("CSV文件READ异常: %s." % e)
            raise Exception("CSV READ ERROR.")

    def write(self, csv_data: list, save_path, insert_str="\t", over_write=True, encode="utf-8", mode="w+"):
        """

        :param csv_data:
        :param save_path:
        :param insert_str:
        :param over_write:
        :param encode:
        :param mode:
        :return:
        """
        if not isinstance(csv_data, list):
            raise Exception("TXT写入数据必须是List<String>格式.")
        if isinstance(csv_data[0], list):
            if isinstance(csv_data[0][0], list) and len(csv_data[0]) == 1:
                csv_data = [sent[0] for sent in csv_data]
                self.logger.warn("写入数据格式不是List<str>, 自动转换为str.")
            else:
                csv_data = [str(sent) for sent in csv_data]
                self.logger.warn("写入数据格式不是List<str>, 强制转换每一个item为str.")

        if over_write:
            self.remove_path(save_path)
        # self.make_path(os.path.splitext(save_path)[0])

        try:
            with open(save_path, mode=mode, encoding=encode) as wf:
                wf.write(insert_str.join(csv_data))
            self.logger.debug("写入TXT数据[%s 条], Path: %s." % (len(csv_data), save_path))
        except Exception as e:
            self.logger.error("TXT文件WRITE异常: %s." % e)
            raise Exception("TXT WRITE Error.")

    def to_excel(self, csv_data, sp_str, sheet_name, save_path, over_write=True):
        """

        :param csv_data:
        :param sp_str:
        :param sheet_name:
        :param save_path:
        :param over_write:
        :return:
        """
        if not (isinstance(csv_data, list) and isinstance(csv_data[0], list)):
            raise Exception("CSV TO EXCEL EXCEPTION: DATA FORMAT ERROR.")

        exce_data = []
        try:
            if over_write:
                self.remove_path(save_path)
            # self.make_path(os.path.splitext(save_path)[0])

            for i, content in enumerate(csv_data):
                if sp_str:
                    content = content.replace("\n", "").split(sp_str)
                exce_data.append(content)
            ExcelHandler().write(exce_data, sheet_name=sheet_name, save_path=save_path)
        except Exception as e:
            self.logger.error("TXT文件WRITE异常: %s." % e)
            raise Exception("TXT WRITE Error.")


class ExcelHandler(BaseHandler):
    def __init__(self):
        super(ExcelHandler, self).__init__()

    def xlwt_write(self, exce_datas, sheet_name, save_path, headers, over_write=False):
        """

        :param exce_datas:
        :param sheet_name:
        :param save_path:
        :param headers:
        :param over_write:
        :return:
        """

    def write(self, exce_datas, sheet_name, save_path, headers, over_write=False):
        """

        :param exce_datas:
        :param sheet_name:
        :param headers:
        :param save_path:
        :param over_write:
        :return:
        """

        if not self.check_suffix(save_path, "xlsx"):
            raise Exception("'to_file' suffix error and must be '.xlsx'.")

        if over_write:
            self.remove_path(save_path)
        # self.make_path(os.path.splitext(save_path)[0])
        if headers:
            exce_datas.insert(0, headers)
        size = len(exce_datas)
        try:
            workbook = openpyxl.Workbook()
            # workbook.title = sheet_name
            ws = workbook.active
            ws.title = sheet_name
            for i in range(0, size):
                if not exce_datas[i]:
                    continue
                for j, value in enumerate(exce_datas[i]):
                    if isinstance(value, str):
                        value = re.sub(ILLEGAL_CHARACTERS_RE, '', str(value))
                    ws.cell(row=i + 1, column=j + 1, value=value)
            workbook.save(save_path)
            self.logger.info("Write [%s] data to Excel[%s] successfully." % (len(exce_datas), save_path))
        except Exception as e:
            raise Exception("EXCEL WRITE Error: %s." % e)

    def append(self, exce_datas, sheet_name, save_path):
        """

        :param exce_datas:
        :param save_path:
        :param sheet_name:
        :return:
        """
        if not self.check_suffix(save_path, "xlsx"):
            raise Exception("'to_file' suffix error and must be '.json'.")
        size = len(exce_datas)
        try:
            workbook = openpyxl.load_workbook(save_path)
            if sheet_name in workbook.get_sheet_names():
                ws = workbook[sheet_name]
            else:
                ws = workbook.create_sheet(sheet_name, 0)
            for i in range(0, size):
                for j in range(0, len(exce_datas[i])):
                    ws.cell(row=i + 1, column=j + 1, value=exce_datas[i][j])
            workbook.save(save_path)
            workbook.close()
        except Exception as e:
            self.logger.error("EXCEL文件APPEND异常: %s." % e)
            raise Exception("EXCEL APPEND Error.")

    def read(self, exce_path, initial_row=0, sheet_names="*", to_dict=False):
        """

        :param exce_path:
        :param initial_row:
        :param sheet_names:
        :param to_dict:
        :return:
        """
        if not self.check_suffix(exce_path, "XLSX"):
            raise Exception("'to_file' suffix error and must be '.json'.")

        def __read__(sheet):
            """
            :param sheet:
            :return:
            """
            sh_data = []
            for row in sheet.iter_rows():
                row_content = [row[i].value for i in range(sheet.max_column)]
                sh_data.append(row_content)
            return sh_data

        ex_datas, ex_dict_datas = [], {}
        try:
            workbook = openpyxl.load_workbook(exce_path, read_only=True, data_only=True)
            _sheet_names = workbook.get_sheet_names()
            if isinstance(sheet_names, int):
                sheet_name = _sheet_names[sheet_names]
                sheet = workbook.get_sheet_by_name(sheet_name)
                ex_datas = __read__(sheet)[initial_row:]

            elif isinstance(sheet_names, list):
                for sheet_name in sheet_names:
                    try:
                        read_state = True
                        if isinstance(sheet_name, str):
                            if sheet_name in _sheet_names:
                                read_state = False
                            else:
                                self.logger.debug("Sheet Name[%s] Not In SheetNames." % (sheet_name))
                        elif isinstance(sheet_name, int):
                            if sheet_name > len(_sheet_names) - 1:
                                self.logger.debug("Sheet Name [%s] In SheetNames." % (sheet_name))
                            else:
                                read_state = False
                        else:
                            self.logger.error("Sheet Name[%s] Error. Must be int or string" % sheet_name)
                        if read_state:
                            sheet = workbook.get_sheet_by_name(sheet_name)
                            _tmp_data = __read__(sheet)
                            tmp_datas = _tmp_data[initial_row:]
                            if to_dict:
                                ex_dict_datas.update({sheet_name: tmp_datas})
                            else:
                                ex_datas += tmp_datas
                            self.logger.debug("> Sheet[%s] 数据[%s条] 读取成功." % (sheet_name, len(_tmp_data)))
                    except Exception as e:
                        self.logger.error("Sheet [%s] 不存在: %s." % (sheet_name, e))
            elif isinstance(sheet_names, str):
                if sheet_names == "*":

                    for sheet_name in _sheet_names:
                        sheet = workbook.get_sheet_by_name(sheet_name)
                        _tmp_data = __read__(sheet)
                        tmp_datas = _tmp_data[initial_row:]
                        if to_dict:
                            ex_dict_datas.update({sheet_name: tmp_datas})
                        else:
                            ex_datas += tmp_datas
                        self.logger.debug("> Sheet[%s] 数据[%s条] 读取成功." % (sheet_name, len(_tmp_data)))
                else:
                    sheet = workbook.get_sheet_by_name(sheet_names)
                    ex_datas = __read__(sheet)[initial_row:]
                    ex_dict_datas.update({sheet_names: ex_datas})
            self.logger.info("Sheet[%s] 数据[%s条] 读取成功." % (sheet_names, len(ex_datas)))
            if to_dict:
                return ex_dict_datas
            else:
                return ex_datas
        except Exception as e:
            self.logger.error("EXCEL文件READ异常: %s." % e)
            raise Exception("EXCEL READ Error.")


class JsonHandler(BaseHandler):
    """
    函数	作用
        json.dumps	对数据进行编码,将python中的字典 转换为 字符串
        json.loads	对数据进行解码,将 字符串 转换为 python中的字典
        json.dump	将dict数据写入json文件中
        json.load	打开json文件，并把字符串转换为python的dict数据

        r	以只读方式打开文件。文件的指针将会放在文件的开头。这是默认模式。
        r+	打开一个文件用于读写。文件指针将会放在文件的开头。
        w+	打开一个文件用于读写。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。
        a	打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。也就是说，新的内容将会被写入到已有内容之后。如果该文件不存在，创建新文件进行写入。
        a+	打开一个文件用于读写。如果该文件已存在，文件指针将会放在文件的结尾。文件打开时会是追加模式。如果该文件不存在，创建新文件用于读写。
    """

    def __init__(self):
        super(JsonHandler, self).__init__()

    def write(self, json_datas, to_file, mode='w', encoding='utf-8', indent=4, ensure_ascii=False, over_write=False):
        """
        写json文件
        :param json_datas:
        :param to_file:
        :param mode:
        :param encoding:
        :param indent:
        :param ensure_ascii:
        :param over_write:
        :return:
        """

        if not self.check_suffix(to_file, "json"):
            raise Exception("'to_file' suffix error and must be '.json'.")

        if over_write:
            self.remove_path(to_file)
        # self.make_path(os.path.splitext(to_file)[0])

        try:
            with open(to_file, mode, encoding=encoding) as fw:
                json.dump(json_datas, fw, indent=indent, ensure_ascii=ensure_ascii)

                # fw.write(json.dumps(json_datas, indent=indent, ensure_ascii=ensure_ascii))
        except Exception as e:
            self.logger.error("JSON文件WRITE异常: %s." % e)
            raise Exception("JSON WRITE Error.")

    def read(self, json_path, mode='r', encoding="utf-8"):
        """

        :param json_path:
        :param mode:
        :param encoding:
        :return:
        """

        try:
            with open(json_path, mode=mode, encoding=encoding) as f:
                json_datas = json.load(f)
            return json_datas
        except Exception as e:
            self.logger.error("JSON文件READ异常: %s." % e)
            raise Exception("JSON READ Error.")


class TXTHandler(BaseHandler):
    """
        r	以只读方式打开文件。文件的指针将会放在文件的开头。这是默认模式。
        r+	打开一个文件用于读写。文件指针将会放在文件的开头。
        w+	打开一个文件用于读写。如果该文件已存在则打开文件，并从开头开始编辑，即原有内容会被删除。如果该文件不存在，创建新文件。
        a	打开一个文件用于追加。如果该文件已存在，文件指针将会放在文件的结尾。也就是说，新的内容将会被写入到已有内容之后。如果该文件不存在，创建新文件进行写入。
        a+	打开一个文件用于读写。如果该文件已存在，文件指针将会放在文件的结尾。文件打开时会是追加模式。如果该文件不存在，创建新文件用于读写。
    """

    def __init__(self):
        super(TXTHandler, self).__init__()

    def read(self, txt_path, space_mark, encode="utf-8", mode="r") -> list:
        """

        :param txt_path:
        :param space_mark:
        :param encode:
        :param mode:
        :return:
        """
        txt_datas = []
        try:
            with open(txt_path, mode=mode, encoding=encode) as rf:
                for content in rf.readlines():
                    try:
                        # content = content.replace("\n", "").strip()
                        if not content:
                            continue
                        if space_mark:
                            content = content.split(space_mark)
                        txt_datas.append(content)
                    except Exception as e:
                        self.logger.warn("异常数据: %s. %s." % (content, e))
        except Exception as e:
            self.logger.error("TXT文件打开异常: %s." % e)
            raise Exception("TXT open Error.")
        return txt_datas

    def write(self, txt_data: list, save_path, insert_str="\n", over_write=False, encode="utf-8", mode="w+"):
        """

        :param txt_data:
        :param save_path:
        :param insert_str:
        :param over_write:
        :param encode:
        :param mode:
        :return:
        """
        if not isinstance(txt_data, list):
            raise Exception("TXT写入数据必须是List<String>格式.")
        # if not isinstance(txt_data[0], str):
        #     txt_data = [sent[0] for sent in txt_data]
        #     self.logger.warn("写入数据格式不是List<str>, 自动转换为str.")

        if over_write:
            self.remove_path(save_path)
        # self.make_path(os.path.splitext(save_path)[0])

        try:
            with open(save_path, mode=mode, encoding=encode) as wf:
                wf.write(insert_str.join(txt_data))
            self.logger.debug("写入TXT数据[%s 条], Path: %s." % (len(txt_data), save_path))
        except Exception as e:
            self.logger.error("TXT文件WRITE异常: %s." % e)
            raise Exception("TXT WRITE Error.")
